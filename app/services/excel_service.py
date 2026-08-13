from io import BytesIO
import re
import unicodedata

from fastapi import UploadFile
from openpyxl import load_workbook


class ExcelService:

    async def ler_excel(
        self,
        arquivo: UploadFile
    ) -> list[dict]:

        conteudo = await arquivo.read()

        if not conteudo:
            raise ValueError(
                "Arquivo Excel vazio ou não informado."
            )

        workbook = load_workbook(
            filename=BytesIO(conteudo),
            data_only=True
        )

        try:

            sheet = workbook.active

            # ========================================================
            # LOCALIZA CABEÇALHO
            # ========================================================

            cabecalho = (
                self._encontrar_linha_cabecalho(
                    sheet
                )
            )

            if cabecalho is None:
                raise ValueError(
                    "Cabeçalho AIRBNB não encontrado na planilha."
                )

            # ========================================================
            # LOCALIZA COLUNAS
            # ========================================================

            colunas = self._encontrar_colunas(
                sheet,
                cabecalho
            )

            if colunas["airbnb"] is None:
                raise ValueError(
                    "Coluna AIRBNB não encontrada."
                )

            # ========================================================
            # COMPETÊNCIA
            # ========================================================

            mes, ano = self._obter_competencia(
                arquivo.filename
            )

            apartamentos = []

            # ========================================================
            # LÊ OS APARTAMENTOS
            # ========================================================

            for linha in sheet.iter_rows(
                min_row=cabecalho + 1
            ):

                numero_apartamento = (
                    self._extrair_numero_apartamento(
                        linha[colunas["airbnb"]].value
                    )
                )

                if not numero_apartamento:
                    continue

                # ====================================================
                # RECEITA BRUTA
                # ====================================================

                receita_bruta = self._obter_valor(
                    linha,
                    colunas["receita_bruta"]
                )

                if receita_bruta <= 0:
                    continue

                # ====================================================
                # CUSTOS
                # ====================================================

                internet = self._obter_valor(
                    linha,
                    colunas["internet"]
                )

                copel = self._obter_valor(
                    linha,
                    colunas["copel"]
                )

                condominio = self._obter_valor(
                    linha,
                    colunas["condominio"]
                )

                limpeza = self._obter_valor(
                    linha,
                    colunas["limpezas"]
                )

                custos_gerais = self._obter_valor(
                    linha,
                    colunas["custos_gerais"]
                )

                comissao = self._obter_valor(
                    linha,
                    colunas["comissao"]
                )

                custos = (
                    internet
                    + copel
                    + condominio
                    + limpeza
                    + custos_gerais
                    + comissao
                )

                # ====================================================
                # RECEITA LÍQUIDA
                # ====================================================

                receita_liquida = self._obter_valor(
                    linha,
                    colunas["rendimento_liquido"]
                )

                if receita_liquida <= 0:
                    receita_liquida = (
                        receita_bruta - custos
                    )

                # ====================================================
                # MÉTRICAS
                # ====================================================

                valor_ap = 240000.0

                ocupacao = 100.0

                locacao_normal = 1000.0

                renda_passiva = receita_liquida

                rentabilidade = (
                    receita_liquida / valor_ap
                ) * 100

                media_mercado = 0.40

                # ====================================================
                # RESULTADO
                # ====================================================

                apartamento = {

                    "apartamento": (
                        f"ED. DUBAI AP "
                        f"{numero_apartamento}"
                    ),

                    "numero_apartamento":
                        numero_apartamento,

                    "mes":
                        mes,

                    "ano":
                        ano,

                    "receita_bruta":
                        receita_bruta,

                    "internet":
                        internet,

                    "copel":
                        copel,

                    "condominio":
                        condominio,

                    "limpeza":
                        limpeza,

                    "reparos":
                        custos_gerais,

                    "administracao":
                        comissao,

                    "custos":
                        custos,

                    "receita_liquida":
                        receita_liquida,

                    "ocupacao":
                        ocupacao,

                    "valor_ap":
                        valor_ap,

                    "locacao_normal":
                        locacao_normal,

                    "renda_passiva":
                        renda_passiva,

                    "rentabilidade":
                        rentabilidade,

                    "media_mercado":
                        media_mercado,
                }

                apartamentos.append(
                    apartamento
                )

            return apartamentos

        finally:

            workbook.close()

    # ================================================================
    # ENCONTRAR CABEÇALHO
    # ================================================================

    def _encontrar_linha_cabecalho(
        self,
        sheet
    ) -> int | None:

        for numero_linha, linha in enumerate(
            sheet.iter_rows(),
            start=1
        ):

            for celula in linha:

                valor = self._normalizar(
                    celula.value
                )

                if valor == "AIRBNB":
                    return numero_linha

        return None

    # ================================================================
    # ENCONTRAR COLUNAS
    # ================================================================

    def _encontrar_colunas(
        self,
        sheet,
        linha_cabecalho: int
    ) -> dict:

        linha = next(
            sheet.iter_rows(
                min_row=linha_cabecalho,
                max_row=linha_cabecalho
            )
        )

        return {

            "airbnb":
                self._encontrar_coluna(
                    linha,
                    "AIRBNB"
                ),

            "receita_bruta":
                self._encontrar_coluna(
                    linha,
                    "REND BRUTO",
                    "REND. BRUTO"
                ),

            "internet":
                self._encontrar_coluna(
                    linha,
                    "INTERNET"
                ),

            "copel":
                self._encontrar_coluna(
                    linha,
                    "COPEL"
                ),

            "condominio":
                self._encontrar_coluna(
                    linha,
                    "CONDOMINIO",
                    "CONDOMÍNIO"
                ),

            "limpezas":
                self._encontrar_coluna(
                    linha,
                    "LIMPEZAS",
                    "LIMPEZA"
                ),

            "custos_gerais":
                self._encontrar_coluna(
                    linha,
                    "CUSTOS G",
                    "CUSTOS G.",
                    "CUSTOS"
                ),

            "comissao":
                self._encontrar_coluna(
                    linha,
                    "COMISSAO",
                    "COMISSÃO"
                ),

            "rendimento_liquido":
                self._encontrar_coluna(
                    linha,
                    "REND LIQUID",
                    "REND. LIQUID",
                    "REND LIQUIDO"
                ),
        }

    # ================================================================
    # ENCONTRAR COLUNA
    # ================================================================

    def _encontrar_coluna(
        self,
        linha,
        *nomes
    ) -> int | None:

        for indice, celula in enumerate(
            linha
        ):

            valor = self._normalizar(
                celula.value
            )

            for nome in nomes:

                nome_normalizado = (
                    self._normalizar(nome)
                )

                if (
                    nome_normalizado
                    in valor
                ):
                    return indice

        return None

    # ================================================================
    # EXTRAIR NÚMERO DO APARTAMENTO
    # ================================================================

    def _extrair_numero_apartamento(
        self,
        valor
    ) -> str | None:

        if valor is None:
            return None

        if isinstance(valor, int):
            return str(valor)

        if isinstance(valor, float):

            if valor.is_integer():
                return str(
                    int(valor)
                )

            return None

        texto = str(
            valor
        ).strip()

        if not texto:
            return None

        resultado = re.search(
            r"\b(\d{3})\b",
            texto
        )

        if resultado:
            return resultado.group(1)

        return None

    # ================================================================
    # OBTER VALOR NUMÉRICO
    # ================================================================

    def _obter_valor(
        self,
        linha,
        indice: int | None
    ) -> float:

        if indice is None:
            return 0.0

        valor = linha[indice].value

        if valor is None:
            return 0.0

        if isinstance(
            valor,
            (int, float)
        ):
            return float(
                valor
            )

        try:

            texto = str(
                valor
            )

            texto = (
                texto
                .replace("R$", "")
                .replace("%", "")
                .replace(".", "")
                .replace(",", ".")
                .strip()
            )

            if not texto:
                return 0.0

            return float(
                texto
            )

        except (
            ValueError,
            TypeError
        ):

            return 0.0

    # ================================================================
    # OBTER COMPETÊNCIA
    # ================================================================

    def _obter_competencia(
        self,
        nome_arquivo: str | None
    ) -> tuple[str, int]:

        if not nome_arquivo:
            raise ValueError(
                "Nome do arquivo não informado."
            )

        resultado = re.search(
            r"(\d{1,2})[-_/](20\d{2})",
            nome_arquivo
        )

        if not resultado:
            raise ValueError(
                "Não foi possível identificar "
                "a competência no nome do arquivo. "
                "Use o formato MM-AAAA."
            )

        mes_numero = int(
            resultado.group(1)
        )

        ano = int(
            resultado.group(2)
        )

        meses = {
            1: "Janeiro",
            2: "Fevereiro",
            3: "Março",
            4: "Abril",
            5: "Maio",
            6: "Junho",
            7: "Julho",
            8: "Agosto",
            9: "Setembro",
            10: "Outubro",
            11: "Novembro",
            12: "Dezembro",
        }

        mes = meses.get(
            mes_numero
        )

        if mes is None:
            raise ValueError(
                f"Mês inválido na competência: "
                f"{mes_numero:02d}"
            )

        return mes, ano

    # ================================================================
    # NORMALIZAR TEXTO
    # ================================================================

    def _normalizar(
        self,
        valor
    ) -> str:

        if valor is None:
            return ""

        texto = str(
            valor
        )

        texto = (
            unicodedata.normalize(
                "NFD",
                texto
            )
        )

        texto = "".join(
            caractere
            for caractere in texto
            if unicodedata.category(
                caractere
            ) != "Mn"
        )

        return (
            texto
            .upper()
            .replace(".", "")
            .replace(":", "")
            .replace("-", " ")
            .strip()
        )